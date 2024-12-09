import telebot
from telebot import types
from telebot.types import ReplyKeyboardRemove
import os
import openpyxl
import openpyxl as op
import config
import time

bot = telebot.TeleBot(token=config.token)
question_and_answer = []
month = [
        {"number": 0, "month": "Январь", "day": 31},
        {"number": 1, "month": "Февраль", "day": 29},
        {"number": 2, "month": "Март", "day": 31},
        {"number": 3, "month": "Апрель", "day": 30},
        {"number": 4, "month": "Май", "day": 31},
        {"number": 5, "month": "Июнь", "day": 30},
        {"number": 6, "month": "Июль", "day": 31},
        {"number": 7, "month": "Август", "day": 31},
        {"number": 8, "month": "Сентябрь", "day": 30},
        {"number": 9, "month": "Октябрь", "day": 31},
        {"number": 10, "month": "Ноябрь", "day": 30},
        {"number": 11, "month": "Декабрь", "day": 31}
    ]


"""
Функции связанные с таблицей
"""


@bot.message_handler(func=lambda message: message.chat.id in config.admin_id, commands=['update'])
def update(message):  # Добавляем аргумент message

    file_name = 'advent.xlsx'
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row

    # Обновляем значения в 5-й колонке
    for i in range(1, max_row + 1):
        sheet.cell(row=i, column=5, value="True")

    # Сохраняем изменения
    wb.save(file_name)

    # Отправляем подтверждение администратору
    bot.send_message(message.chat.id, "Все задания обновлены.")
    wb.close()


def create_buttons_podpiska():
    marcup = types.InlineKeyboardMarkup(row_width=True)
    marcup.add(
        types.InlineKeyboardButton(text="канал", url="https://t.me/goodday_marketing"),
        types.InlineKeyboardButton(text="проверить", callback_data="check")
    )
    return marcup


@bot.message_handler(func=lambda message: message.chat.id in config.admin_id, commands=['xlsx_file'])
def send_file(message):
    bot.send_document(message.chat.id, document=open("users.xlsx", "rb"))
    bot.send_document(message.chat.id, document=open("advent.xlsx", "rb"))


def open_question_xlsx():
    global question_and_answer
    file_name = 'advent.xlsx'
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    for i in range(1, max_row+1):
        question_and_answer.append({"question": sheet.cell(row=i, column=1).value,
                                    "answer": sheet.cell(row=i, column=2).value,
                                    "answer1": sheet.cell(row=i, column=3).value,
                                    "format": sheet.cell(row=i, column=4).value,
                                    "TF": sheet.cell(row=i, column=5).value,
                                    "number": str(i - 1)})


open_question_xlsx()


def create_excel_file():
    file_name = "users.xlsx"
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Users'
        qwe = ['ID', 'Username', 'First_name', 'fio', 'fm', 'number', "hb"]
        for i in range(1, 32):
            qwe.append(f'quiz {i}')
        sheet.append(qwe)

        workbook.save(file_name)
        print(f"file {file_name} create")


create_excel_file()


def save_fio(fio1, message):
    file_name = 'users.xlsx'
    if os.path.exists(file_name):
        workbook = op.load_workbook(file_name, data_only=True)
        sheet = workbook.active
        max_row = sheet.max_row + 1
        # Добавляем данные явно
        username = message.chat.username
        if not username:
            username = "None"
        sheet.cell(row=max_row, column=1, value=message.chat.id)
        sheet.cell(row=max_row, column=2, value=username)
        sheet.cell(row=max_row, column=3, value=message.chat.first_name)
        sheet.cell(row=max_row, column=4, value=fio1)
        for i in range(8, 39):
            sheet.cell(row=max_row, column=i, value="False")
        try:
            workbook.save(file_name)
        finally:
            workbook.close()
    else:
        create_excel_file()
        save_fio(fio1, message)  # Исправлено на правильное использование fio1


"""
функции связанные с обработкой сообщений
"""


def message_photo(message, data):
    try:
        photo = message.photo[-1]
        file_info = bot.get_file(photo.file_id)
        downloaded_file = bot.download_file(file_info.file_path)  # Скачиваем файл

        # Создаем клавиатуру для администратора
        marcup = types.InlineKeyboardMarkup(row_width=2)
        marcup.add(
            types.InlineKeyboardButton("прошел", callback_data=f"True|{data}|{message.chat.id}"),
            types.InlineKeyboardButton("Не прошел", callback_data=f"False|{data}|{message.chat.id}")
        )

        # Отправляем фото администраторам
        for i in config.admin_id:
            bot.send_photo(
                i,
                downloaded_file,
                reply_markup=marcup,
                caption=f"Ответ пользователя @{message.chat.username}: \n{question_and_answer[int(data)]["question"]}"
            )

        # Работа с файлом Excel
        file_name = 'users.xlsx'
        if not os.path.exists(file_name):
            raise FileNotFoundError(f"Файл {file_name} не найден.")

        workbook = op.load_workbook(file_name, data_only=True)
        sheet = workbook.active

        # Расчёт индекса столбца
        column_index = 8 + int(data)  # Индекс столбца для сохранения

        # Поиск строки с нужным chat_id
        for row in sheet.iter_rows(min_row=2):  # Обрабатываем строки начиная со 2-й
            if row[0].value == message.chat.id:  # Если chat_id совпадает
                # Обновляем ячейку с фото
                sheet.cell(row=row[0].row, column=column_index, value="Фото сохранено")
                break

        workbook.save(file_name)
        workbook.close()

    except Exception as e:
        bot.send_message(message.chat.id, f"Произошла ошибка: {str(e)}")


def message_sms(message, data):
    try:
        sms1 = message.text
        print(data)  # Проверяем, что приходит в data (ожидаем число)

        # Создаем клавиатуру для администратора
        marcup = types.InlineKeyboardMarkup(row_width=2)
        marcup.add(
            types.InlineKeyboardButton("прошел", callback_data=f"True|{data}|{message.chat.id}"),
            types.InlineKeyboardButton("Не прошел", callback_data=f"False|{data}|{message.chat.id}")
        )

        # Отправляем сообщение администраторам
        for i in config.admin_id:
            bot.send_message(
                i,
                f"{question_and_answer[int(data)]['question']}\nОтвет пользователя: @{message.chat.username} \n{sms1}",
                reply_markup=marcup
            )

        # Работа с файлом Excel
        file_name = 'users.xlsx'
        if not os.path.exists(file_name):
            raise FileNotFoundError(f"Файл {file_name} не найден.")

        workbook = op.load_workbook(file_name, data_only=True)
        sheet = workbook.active

        # Расчёт индекса столбца
        column_index = 8 + int(data)  # Начинаем с 8-го столбца

        # Поиск строки с нужным chat_id
        for row in sheet.iter_rows(min_row=2):  # Обрабатываем строки начиная со 2-й
            if row[0].value == message.chat.id:  # Если chat_id совпадает
                # Обновляем ячейку с текстом
                sheet.cell(row=row[0].row, column=column_index, value=sms1)
                break

        workbook.save(file_name)
        workbook.close()

    except Exception as e:
        bot.send_message(message.chat.id, f"Произошла ошибка: {str(e)}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("True|"))
def answer_true(call):
    try:
        data = call.data.split("|")
        print(data)

        # Отправляем сообщение пользователю
        bot.send_message(
            int(data[2]),
            f"Поздравляю, вот ваш выигрыш: {question_and_answer[int(data[1])]['answer']}:\n"
            f"{question_and_answer[int(data[1])]['answer1']}")

    except Exception as e:
        # Логирование ошибки
        bot.send_message(call.message.chat.id, f"Произошла ошибка: {str(e)}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("False|"))
def answer_false(call):
    data = call.data.split("|")

    # Работа с файлом Excel
    file_name = 'users.xlsx'
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"Файл {file_name} не найден.")

    workbook = op.load_workbook(file_name, data_only=True)
    sheet = workbook.active

    # Расчёт индекса столбца
    column_index = 8 + int(data[1])  # Начинаем с 8-го столбца

    # Поиск строки с нужным chat_id
    for row in sheet.iter_rows(min_row=2):  # Обрабатываем строки начиная со 2-й
        if row[0].value == int(data[2]):  # Если chat_id совпадает
            # Обновляем ячейку с текстом
            sheet.cell(row=row[0].row, column=column_index, value="False")
            break

    workbook.save(file_name)
    workbook.close()
    bot.send_message(int(data[2]),
                     f"Вы не прошли проверку пришлите корректный ответ", reply_markup=create_buttons_quiz(data[1]))


@bot.message_handler(func=lambda message: message.chat.id in config.admin_id, commands=['start_quiz'])
def quiz(message):
    file_name = 'advent.xlsx'
    user_file = 'users.xlsx'

    # Проверяем наличие файлов
    if not os.path.exists(file_name) or not os.path.exists(user_file):
        bot.send_message(message.chat.id, "Отсутствуют необходимые файлы.")
        return

    # Загружаем вопросы
    workbook = op.load_workbook(file_name, data_only=True)
    sheet = workbook.active

    # Загружаем пользователей
    user_workbook = op.load_workbook(user_file, data_only=True)
    user_sheet = user_workbook.active

    # Цикл по вопросам
    for i in range(len(question_and_answer)):
        if question_and_answer[i]["TF"] == "True":  # Если задание активно
            question_and_answer[i]["TF"] = "False"  # Помечаем задание как выполненное

            # Обновляем статус задания в Excel
            row_index = i + 1
            column_index = 5  # Предполагаем, что статус в 5-й колонке
            sheet.cell(row=row_index, column=column_index, value="False")
            workbook.save(file_name)

            # Цикл по пользователям
            for row in user_sheet.iter_rows(min_row=2,
                                            max_row=user_sheet.max_row,
                                            min_col=1,
                                            max_col=7,
                                            values_only=True):
                user_id = row[0]
                if not user_id:
                    continue

                # Проверяем, проходил ли пользователь задание

                # Отправляем задание пользователю
                bot.send_message(
                    user_id,
                    f"Привет, {row[3]}!\n{question_and_answer[i]['question']}",
                    reply_markup=create_buttons_quiz(i)
                )

            # Сохраняем изменения в файле пользователей
            user_workbook.save(user_file)
            user_workbook.close()
            break


@bot.callback_query_handler(func=lambda call: call.data.startswith("quiz1|"))
def question(call):
    global question_and_answer
    data = call.data.split("|")
    marcup = types.InlineKeyboardMarkup(row_width=2)
    print(data)
    for i in range(len(question_and_answer)):
        # Проверяем, соответствует ли вопрос выбранному
        if question_and_answer[i]["number"] == data[1]:
            # Добавляем кнопку для ответа
            marcup.add(
                types.InlineKeyboardButton("ответить", callback_data=f"quiz|{question_and_answer[i]['number']}"),
                types.InlineKeyboardButton("Назад", callback_data="return")
            )
            # Отправляем вопрос с кнопкой
            bot.send_message(call.message.chat.id, question_and_answer[i]["question"], reply_markup=marcup)
            break  # Прекращаем цикл после нахождения подходящего вопроса


@bot.callback_query_handler(func=lambda call: call.data.startswith("quiz|"))
def answer(call):
    data = call.data.split("|")
    file_name = 'users.xlsx'
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    qwe = True
    for i in range(2, max_row + 1):
        if (int(sheet.cell(row=i, column=1).value) == call.message.chat.id and
                sheet.cell(row=i, column=8 + int(data[1])).value == "False"):
            qwe = False
    if not qwe:
        if question_and_answer[int(data[1])]["format"] == "Текст":
            text = bot.send_message(call.message.chat.id, "пришлите сообщение ответ")
            bot.register_next_step_handler(text, lambda message: message_sms(message, data[1]))
        else:
            foto = bot.send_message(call.message.chat.id, "пришлите фото в качестве ответа")
            bot.register_next_step_handler(foto, lambda message: message_photo(message, data[1]))
    else:
        bot.send_message(call.message.chat.id, "вы уже прошли этот квест")
        return


@bot.callback_query_handler(func=lambda call: call.data.startswith("day|"))
def save_month_data(call):
    # Разделяем данные callback на части
    month_data = call.data.split("|")

    # Убедимся, что формат данных корректен
    if len(month_data) != 2 or not month_data[1]:
        bot.answer_callback_query(call.id, "Ошибка: некорректные данные.")
        return

    # Удаляем сообщение с кнопками
    try:
        bot.delete_message(call.message.chat.id, call.message.id)
    except Exception as e:
        print(f"Ошибка удаления сообщения: {e}")
    formatted_date = month_data[1]  # Дата уже отформатирована (например, 01.04)

    file_name = 'users.xlsx'
    if os.path.exists(file_name):
        workbook = op.load_workbook(file_name, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=6):  # Указываем max_col=7, чтобы обработать 7 колонок
            if row[0].value == call.message.chat.id:  # Проверяем, что ID совпадает
                # Сохраняем номер в 7-й колонке
                sheet.cell(row=row[0].row, column=7, value=formatted_date)
                break
        for i in config.admin_id:
            for row in sheet.iter_rows(min_row=2, max_row=6):
                if row[0].value == call.message.chat.id:
                    bot.send_message(i,
                                     f"Новый пользователь:\n"
                                     f"ID: {row[0].value},\n"
                                     f"Username: {row[1].value},\n"
                                     f"First name: {row[2].value},\n"
                                     f"Имя: {row[3].value},\n"
                                     f"пол: {row[4].value},\n"
                                     f"номер телефона: +{row[5].value},\n"
                                     f"Дата Рождения: {row[6].value},\n")

        workbook.save(file_name)
        workbook.close()
    # Отправляем подтверждение пользователю
    bot.send_message(
        call.message.chat.id,
        f"Отлично! Теперь мы знаем когда тебя поздравлять 🙃🤗\n"
        f"Для продолжения работы подпишитесь на канал", reply_markup=create_buttons_podpiska()
    )


@bot.callback_query_handler(func=lambda call: call.data.startswith("check"))
def check(call):
    status = ['creator', 'administration', 'member']
    for i in status:
        if (i == bot.get_chat_member(chat_id="-1001929329164", user_id=call.message.chat.id).status
                or call.message.chat.id in config.admin_id):
            bot.send_message(call.message.chat.id, config.text_5)
            break
    else:
        bot.send_message(call.message.chat.id, "Подпишитесь пожалуйста", reply_markup=create_buttons_podpiska())


@bot.callback_query_handler(func=lambda call: call.data.startswith("month|"))
def save_month(call):
    print(call.data)
    bot.delete_message(call.message.chat.id, call.message.id)
    bot.send_message(call.message.chat.id, f"Отлично, теперь выберите день рождения",
                     reply_markup=create_buttons_day(call.data))


@bot.callback_query_handler(func=lambda call: call.data.startswith("start1|"))
def message_2(call):
    user_id = call.message.chat.id
    bot.delete_message(call.message.chat.id, call.message.id)
    bot.send_message(user_id, config.text_2, reply_markup=create_buttons_start2())


@bot.callback_query_handler(func=lambda call: call.data.startswith("start2|"))
def message_3(call):
    bot.delete_message(call.message.chat.id, call.message.id)
    user_id = call.message.chat.id
    bot.send_message(user_id, config.text_3, reply_markup=create_buttons_start3())


@bot.callback_query_handler(func=lambda call: call.data.startswith("start3|"))
def message_4(call):
    bot.delete_message(call.message.chat.id, call.message.id)
    user_id = call.message.chat.id
    message = bot.send_message(user_id, config.text_4, reply_markup=ReplyKeyboardRemove())
    bot.register_next_step_handler(message, fio)


@bot.callback_query_handler(func=lambda call: call.data.startswith("male|"))
def save_female(call):
    file_name = 'users.xlsx'
    male = call.data.split("|")
    if os.path.exists(file_name):
        workbook = op.load_workbook(file_name, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=6):  # Обрабатываем строки с 2-й
            if row[0].value == call.message.chat.id:  # Проверяем, что ID совпадает
                if row[5] is not None:  # Проверяем, что ячейка существует
                    sheet.cell(row=row[0].row, column=5, value=male[1])  # Обновляем 6-ю колонку
                else:
                    sheet.cell(row=row[0].row, column=5, value=male[1])  # Добавляем данные
                break

        workbook.save(file_name)
        workbook.close()
        bot.delete_message(call.message.chat.id, call.message.id)
        bot.send_message(call.message.chat.id,
                         "Пожалуйста укажите номер телефон для получения подарков(достаточно нажать на кнопку)",
                         reply_markup=create_buttons_phone_number())
    else:
        create_excel_file()
        save_female(call)  # Рекурсивный вызов функции


@bot.message_handler(content_types=['contact'])
def save_phone(message):
    file_name = 'users.xlsx'
    bot.send_message(message.chat.id, "Спасибо!",
                     reply_markup=ReplyKeyboardRemove())
    if os.path.exists(file_name):
        workbook = op.load_workbook(file_name, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=6):  # Указываем max_col=6 для обработки 6 колонок
            if row[0].value == message.chat.id:  # Проверяем, что ID совпадает
                # Сохраняем номер в 6-й колонке
                sheet.cell(row=row[0].row, column=6, value=message.contact.phone_number)
                break

        workbook.save(file_name)
        workbook.close()
        # Удаляем клавиатуру и отправляем сообщение
        bot.send_message(message.chat.id, "Выберете месяц рождения",
                         reply_markup=create_buttons_month())
    else:
        create_excel_file()
        save_phone(message)

    # Удаляем клавиатуру после обработки


def fio(message):
    text = message.text
    bot.send_message(message.chat.id, f"{text}, приятно познакомиться\n теперь укажите свой пол",
                     reply_markup=create_button_female_male())
    save_fio(text, message)


@bot.message_handler(commands=['start', 'help'])
def start(message):
    user_id = message.chat.id
    file_name = 'users.xlsx'
    if not os.path.exists(file_name):
        create_excel_file()

    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    user_id1 = True
    for i in range(1, max_row + 1):
        if sheet.cell(row=i, column=1).value == user_id:
            user_id1 = False
    if user_id1:
        bot.send_message(user_id, config.text_1, reply_markup=create_buttons_start1())
    else:
        bot.send_message(user_id, "Скоро появятся задания")


@bot.message_handler(commands=['technical_support'])
def support(message):
    sms2 = bot.send_message(message.chat.id, "Пожалуйста, опишите проблему")
    bot.register_next_step_handler(sms2, support_answer)


def support_answer(message):
    marcup = types.InlineKeyboardMarkup(row_width=1)
    marcup.add(types.InlineKeyboardButton(text="Ответить", callback_data=f"support_answer|{message.chat.id}"))
    for i in config.admin_id:
        bot.send_message(i, message.text, reply_markup=marcup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("support_answer|"))
def answer_admin(call):
    data = call.data.split("|")
    sms3 = bot.send_message(call.message.chat.id, "введите ответ")
    bot.register_next_step_handler(sms3, lambda message: answer_admin_message(message, data[1]))


def answer_admin_message(message, data):
    bot.send_message(int(data), f"Сообщение от администрации:\n{message.text}")


@bot.message_handler(func=lambda message: message.chat.id in config.admin_id, commands=['notification'])
def notification(message):
    sms4 = bot.send_message(message.chat.id, "Напишите новость")
    bot.register_next_step_handler(sms4, notification_sms)


def notification_sms(message):
    file_name = 'users.xlsx'
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row
    for i in range(2, max_row+1):
        bot.send_message(int(sheet.cell(row=i, column=1).value), f"{message.text}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("return"))
def task_return(call):
    task(call.message)


@bot.message_handler(commands=['task'])
def task(message):
    marcup = types.InlineKeyboardMarkup(row_width=True)
    asd = False
    for i in range(0, len(question_and_answer)):
        if question_and_answer[i]["TF"] == "False":
            asd = "True"
            marcup.add(types.InlineKeyboardButton(text=question_and_answer[i]["question"],
                                                  callback_data=f"quiz1|{question_and_answer[i]["number"]}"))
    if asd == "True":
        bot.send_message(message.chat.id, "выберете интересующее задание", reply_markup=marcup)
    else:
        bot.send_message(message.chat.id, "Пока нет заданий")


@bot.message_handler(commands=['connect'])
def connect(message):
    bot.send_message(message.chat.id, "Стать спонсором @ulyana_goodday")


"""Кнопки"""


def create_buttons_quiz(questions):
    marcup = types.InlineKeyboardMarkup(row_width=2)
    marcup.add(
        types.InlineKeyboardButton(text="Отправить ответ", callback_data=f"quiz|{questions}")
    )
    return marcup


def create_buttons_month():
    global month
    marcup = types.InlineKeyboardMarkup(row_width=3)
    for i in range(0, len(month), 3):
        marcup.add(
            types.InlineKeyboardButton(month[i]["month"], callback_data=f"month|{month[i]["number"]}"),
            types.InlineKeyboardButton(month[i+1]["month"], callback_data=f"month|{month[i+1]["number"]}"),
            types.InlineKeyboardButton(month[i+2]["month"], callback_data=f"month|{month[i+2]["number"]}")
        )
    return marcup


def create_buttons_day(data):
    global month

    # Проверяем, что данные переданы корректно
    try:
        number = data.split("|")
        selected_month = int(number[1])  # Номер месяца
    except (IndexError, ValueError):
        raise ValueError("Неверный формат данных. Ожидается строка формата 'month|MM'.")

    # Проверяем, что номер месяца существует в данных month
    if selected_month < 0 or selected_month >= len(month):
        raise ValueError(f"Месяц {selected_month} не найден в данных 'month'.")

    # Извлекаем количество дней в месяце
    days_in_month = month[selected_month]["day"]

    # Создаём клавиатуру
    markup = types.InlineKeyboardMarkup(row_width=7)

    # Формируем кнопки по неделям
    for i in range(1, days_in_month + 1, 7):
        buttons = []
        for j in range(i, min(i + 7, days_in_month + 1)):
            # Форматируем день и месяц с ведущими нулями
            formatted_day = f"{j:02}"  # День с ведущим нулем
            formatted_month = f"{selected_month + 1:02}"  # Месяц с ведущим нулем (индекс + 1)

            buttons.append(types.InlineKeyboardButton(
                text=f"{j}",
                callback_data=f"day|{formatted_day}.{formatted_month}"  # Формируем callback_data
            ))
        markup.add(*buttons)

    return markup


def create_buttons_start1():
    marcup = types.InlineKeyboardMarkup(row_width=1)
    marcup.add(
        types.InlineKeyboardButton("Да, хочу попробовать!", callback_data="start1|1"),
        types.InlineKeyboardButton("Звучит очень интересно!", callback_data="start1|2"),
        types.InlineKeyboardButton("Никогда не участвовал(а) в таком, хочу попробовать!", callback_data="start1|3")
    )
    return marcup


def create_buttons_phone_number():
    marcup = types.ReplyKeyboardMarkup(one_time_keyboard=True, row_width=1)
    marcup.add(types.KeyboardButton("Поделиться номером телефона", request_contact=True))
    return marcup


def create_buttons_start2():
    marcup = types.InlineKeyboardMarkup(row_width=1)
    marcup.add(
        types.InlineKeyboardButton("Курсы и чек-листы – это круто!", callback_data="start2|1"),
        types.InlineKeyboardButton("Всё, что поможет вдохновиться", callback_data="start2|2"),
        types.InlineKeyboardButton("Не могу дождаться первого задания!", callback_data="start2|3")
    )
    return marcup


def create_buttons_start3():
    marcup = types.InlineKeyboardMarkup(row_width=1)
    marcup.add(
        types.InlineKeyboardButton("Определенно, да!", callback_data="start3|1"),
        types.InlineKeyboardButton("Давай начнем!", callback_data="start3|2"),
        types.InlineKeyboardButton("Люблю такие форматы, готов(а) участвовать!", callback_data="start3|3")
    )
    return marcup


def create_button_female_male():
    marcup = types.InlineKeyboardMarkup(row_width=2)
    marcup.add(
        types.InlineKeyboardButton("Мужской", callback_data="male|Мужской"),
        types.InlineKeyboardButton("Женский", callback_data="male|Женский"))
    return marcup


def run_bot():
    while True:
        try:
            bot.polling(non_stop=True, timeout=60)
        except Exception as e:
            print(f"Ошибка: {e}")
            time.sleep(5)


run_bot()
